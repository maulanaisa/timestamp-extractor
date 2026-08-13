import os
import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image as pyxlImage
from tqdm import tqdm
from PIL import Image, ImageOps

from utils import delete_directory, create_directory, detect_closest_coordinate, coor_csv_to_dict

class Pohon:
    def __init__(self, settings):
        self.wb = Workbook()
        self.ws = self.wb.active

        self.openai_api_path = "output_openai_api.csv"
        self.images_path = settings["IMAGES_PATH"]
        self.export_path = settings["EXPORT_PATH"]
        self.image_dimension = settings["IMAGE_DIMENSION"]
        self.location_coordinates = settings["LOCATION_COORDINATES"]
        self.image_quality = settings["IMAGE_QUALITY"]

    def images_with_coordinates(self):
        if os.path.exists(self.openai_api_path):
            row = 1
            temp_folder_name = os.path.join(self.images_path, "temp")
            create_directory(temp_folder_name)

            with open(self.openai_api_path, 'r', newline='') as csvfile:
                reader = csv.reader(csvfile)

                # Iterate over each row in the CSV file
                for csv_row in tqdm(reader):
                    file_name = csv_row[0]
                    latitude = csv_row[1]
                    longitude = csv_row[2]

                    img_path = os.path.join(self.images_path, file_name)
                    temp_img_path = os.path.join(temp_folder_name, f"{os.path.splitext(file_name)[0]}.jpg")   # Temporary path to store fixed images
                    try:
                        with Image.open(img_path) as img:
                            img_fixed = ImageOps.exif_transpose(img)    # Fix image orientation and save to temporary path
                            img_fixed = img_fixed.convert('RGB')
                            img_fixed.save(temp_img_path, "JPEG", quality=self.image_quality, optimize=True)

                        img_pyxl = pyxlImage(temp_img_path)

                        img_pyxl.height = self.image_dimension["height"]
                        img_pyxl.width = self.image_dimension["width"]
                        
                        # If location .csv file exists
                        if os.path.exists(self.location_coordinates):
                            coor_dict = coor_csv_to_dict(self.location_coordinates)    
                            location = detect_closest_coordinate(coor_dict, (latitude, longitude))
                        else:
                            location = ""

                        self.ws[f"A{row}"] = row
                        self.ws[f"B{row}"] = location
                        self.ws[f"C{row}"] = f"{latitude}, {longitude}"
                        self.ws.add_image(img_pyxl, f"D{row}")  

                    except:

                        self.ws[f"A{row}"] = row
                        self.ws[f"B{row}"] = location
                        self.ws[f"C{row}"] = f"{latitude}, {longitude}"

                    row += 1

            self.wb.save(self.export_path)
            print(f"Workbook {self.export_path} is created successfully.")

            delete_directory(temp_folder_name)
                                    
        else:
            raise FileNotFoundError("output_openai_api.csv not found!, please generate first.")
